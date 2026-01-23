"""
데이터베이스 연결 및 CRUD 함수
PostgreSQL (Supabase) 전용 - 최적화 버전
"""
import os
import re
import streamlit as st
import psycopg2
import openpyxl
from psycopg2.extras import RealDictCursor
from datetime import datetime
from typing import List, Dict

# ---------------------------------------------------------
# 1. DB 연결 및 설정 (캐싱 적용)
# ---------------------------------------------------------

def validate_connection(conn):
    """연결 유효성 검사"""
    try:
        return conn.closed == 0
    except:
        return False

def clear_cache():
    """데이터 변경(CUD) 시 캐시 무효화"""
    st.cache_data.clear()

@st.cache_resource(ttl=3600, validate=validate_connection)
def get_connection():
    """DB 연결 (Supabase)"""
    try:
        # 1. secrets.toml의 DATABASE_URL 우선 사용
        db_url = st.secrets.get("DATABASE_URL")
        if db_url:
            return psycopg2.connect(db_url)
        
        # 2. [supabase] 섹션 사용 (Legacy)
        if "supabase" in st.secrets:
            return psycopg2.connect(st.secrets["supabase"]["db_url"])
            
        raise Exception("secrets.toml에 DATABASE_URL이 없습니다.")
    except Exception as e:
        st.error(f"DB 연결 실패: {e}")
        raise e

def get_cursor(conn):
    return conn.cursor(cursor_factory=RealDictCursor)

@st.cache_resource 
def init_db():
    """DB 테이블 초기화 (최초 1회만 실행)"""
    conn = get_connection()
    with conn.cursor() as cursor:
        queries = [
            """
            CREATE TABLE IF NOT EXISTS participants (
                name TEXT NOT NULL,
                birth_date TEXT NOT NULL,
                gender TEXT NOT NULL,
                nickname TEXT, phone TEXT, location TEXT, job TEXT, mbti TEXT, 
                intro TEXT, signup_route TEXT, first_visit_date TEXT, memo TEXT,
                PRIMARY KEY (name, birth_date)
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS sessions (
                session_id SERIAL PRIMARY KEY,
                session_date TEXT NOT NULL, session_time TEXT, 
                theme TEXT, host TEXT, status TEXT DEFAULT '준비중'
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS attendance (
                attendance_id SERIAL PRIMARY KEY,
                participant_name TEXT NOT NULL, participant_birth TEXT NOT NULL,
                session_id INTEGER NOT NULL, attended BOOLEAN DEFAULT TRUE, payment_status TEXT,
                FOREIGN KEY (participant_name, participant_birth) REFERENCES participants(name, birth_date),
                FOREIGN KEY (session_id) REFERENCES sessions(session_id)
            )
            """
        ]
        for query in queries:
            cursor.execute(query)
        conn.commit()
    print("✅ DB 초기화 완료! (최초 1회 실행됨)")

# ---------------------------------------------------------
# 2. 데이터 생성 (INSERT) - 실행 후 clear_cache()
# ---------------------------------------------------------

def add_participant(name: str, birth_date: str, gender: str, 
                   job: str = "", mbti: str = "", phone: str = "", 
                   location: str = "", signup_route: str = "", memo: str = ""):
    """참가자 추가"""
    conn = get_connection()
    try:
        with get_cursor(conn) as cursor:
            cursor.execute("""
                INSERT INTO participants 
                (name, birth_date, gender, job, mbti, phone, location, signup_route, first_visit_date, memo)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (name, birth_date) DO NOTHING
            """, (name, birth_date, gender, job, mbti, phone, location, signup_route, 
                  datetime.now().strftime("%Y-%m-%d"), memo))
            conn.commit()
            clear_cache()
            print(f"✅ {name} 추가 완료!")
            return True
    except Exception as e:
        conn.rollback()
        print(f"❌ 추가 실패: {e}")
        return False

def create_session(session_date, session_time, theme, host=""):
    """회차 생성"""
    conn = get_connection()
    try:
        with get_cursor(conn) as cursor:
            cursor.execute("""
                INSERT INTO sessions (session_date, session_time, theme, host)
                VALUES (%s, %s, %s, %s)
                RETURNING session_id
            """, (session_date, session_time, theme, host))
            session_id = cursor.fetchone()['session_id']
            conn.commit()
            clear_cache()
            print(f"✅ 회차 생성 완료! ID: {session_id}")
            return session_id
    except Exception as e:
        conn.rollback()
        raise e

def add_attendance(session_id: int, participant_name: str, participant_birth: str):
    """회차에 참가자 추가"""
    conn = get_connection()
    try:
        with get_cursor(conn) as cursor:
            cursor.execute("""
                INSERT INTO attendance (session_id, participant_name, participant_birth)
                VALUES (%s, %s, %s)
                ON CONFLICT DO NOTHING
            """, (session_id, participant_name, participant_birth))
            conn.commit()
            clear_cache()
            print(f"✅ 출석 추가 완료: {participant_name}")
    except Exception as e:
        conn.rollback()
        raise e

# ---------------------------------------------------------
# 3. 데이터 조회 (SELECT) - @st.cache_data 적용
# ---------------------------------------------------------

# 💡 _cache_version=0 을 파라미터에 추가해서 app.py와의 충돌을 방지합니다.

@st.cache_data(ttl=600)
def get_all_participants(_cache_version=0) -> List[Dict]:
    """모든 참가자 조회"""
    conn = get_connection()
    with get_cursor(conn) as cursor:
        cursor.execute("""
            SELECT name, birth_date, gender, job, mbti, phone, location, signup_route, first_visit_date, memo
            FROM participants ORDER BY name
        """)
        return [dict(row) for row in cursor.fetchall()]

@st.cache_data(ttl=600)
def get_all_sessions(_cache_version=0) -> List[Dict]:
    """모든 회차 조회"""
    conn = get_connection()
    with get_cursor(conn) as cursor:
        cursor.execute("""
            SELECT session_id, session_date, session_time, theme, host, status
            FROM sessions ORDER BY session_date DESC, session_time DESC
        """)
        return [dict(row) for row in cursor.fetchall()]

@st.cache_data(ttl=600)
def get_session_participants(session_id: int, _cache_version=0) -> List[Dict]:
    """특정 회차의 참가자 목록 (방문 횟수 + 메모 포함)"""
    conn = get_connection()
    with get_cursor(conn) as cursor:
        cursor.execute("""
            SELECT p.name, p.birth_date, p.gender, p.job, p.mbti, p.phone,
                   p.location, p.signup_route, p.memo,  -- 🔥 [수정] 메모 컬럼 추가!
                   a.attendance_id, a.payment_status,
                   (SELECT COUNT(*) FROM attendance a2 
                    WHERE a2.participant_name = p.name 
                    AND a2.participant_birth = p.birth_date) as visit_count
            FROM attendance a
            JOIN participants p ON a.participant_name = p.name 
                                AND a.participant_birth = p.birth_date
            WHERE a.session_id = %s
        """, (session_id,))
        return [dict(row) for row in cursor.fetchall()]

# ---------------------------------------------------------
# 4. 고급 로직 (Logic) - N+1 문제 해결 및 최적화
# ---------------------------------------------------------

@st.cache_data(ttl=600)
def check_duplicate_meetings(session_id: int, _cache_version=0) -> List[Dict]:
    """중복 만남 확인 (Bulk Fetching 최적화)"""
    conn = get_connection()
    current_participants = get_session_participants(session_id) # 여기서는 내부 호출이라 인자 없음
    
    if len(current_participants) < 2: return []

    targets = [(p['name'], p['birth_date']) for p in current_participants]
    
    with get_cursor(conn) as cursor:
        conditions = ["(participant_name = %s AND participant_birth = %s)"] * len(targets)
        where_clause = " OR ".join(conditions)
        params = [val for t in targets for val in t]
        params.append(session_id)

        query = f"""
            SELECT participant_name, participant_birth, s.session_date
            FROM attendance a
            JOIN sessions s ON a.session_id = s.session_id
            WHERE ({where_clause}) AND a.session_id != %s
        """
        cursor.execute(query, tuple(params))
        history = cursor.fetchall()

    history_map = {}
    for row in history:
        date = row['session_date']
        person = (row['participant_name'], row['participant_birth'])
        if date not in history_map: history_map[date] = []
        history_map[date].append(person)
    
    duplicates = []
    for date, people in history_map.items():
        if len(people) < 2: continue
        for i in range(len(people)):
            for j in range(i + 1, len(people)):
                p1, p2 = people[i], people[j]
                if p1 in targets and p2 in targets:
                    if p1 > p2: p1, p2 = p2, p1
                    
                    found = False
                    for d in duplicates:
                        if d['person1'] == p1[0] and d['person2'] == p2[0]:
                            if date not in d['session_dates']:
                                d['session_dates'].append(date)
                            found = True
                            break
                    
                    if not found:
                        duplicates.append({
                            'person1': p1[0], 'person1_birth': p1[1],
                            'person2': p2[0], 'person2_birth': p2[1],
                            'session_dates': [date]
                        })
    
    for d in duplicates: d['session_dates'].sort()
    return duplicates

@st.cache_data(ttl=600)
def get_participant_detail(name: str, birth_date: str, _cache_version=0) -> Dict:
    """참가자 상세 정보 (이력 포함)"""
    conn = get_connection()
    with get_cursor(conn) as cursor:
        cursor.execute("SELECT * FROM participants WHERE name = %s AND birth_date = %s", (name, birth_date))
        row = cursor.fetchone()
        if not row: return {}
        participant = dict(row)
        
        cursor.execute("""
            SELECT s.session_id, s.session_date, s.session_time, s.theme
            FROM attendance a
            JOIN sessions s ON a.session_id = s.session_id
            WHERE a.participant_name = %s AND a.participant_birth = %s
            ORDER BY s.session_date DESC
        """, (name, birth_date))
        participant['visit_history'] = [dict(r) for r in cursor.fetchall()]
        participant['visit_count'] = len(participant['visit_history'])
        
        for visit in participant['visit_history']:
            cursor.execute("""
                SELECT p.name, p.gender
                FROM attendance a
                JOIN participants p ON a.participant_name = p.name AND a.participant_birth = p.birth_date
                WHERE a.session_id = %s AND NOT (p.name = %s AND p.birth_date = %s)
            """, (visit['session_id'], name, birth_date))
            visit['met_people'] = [dict(r) for r in cursor.fetchall()]
            
    return participant

@st.cache_data(ttl=600)
def get_recommendations(session_id: int, gender: str, age_min: int = None, age_max: int = None, mbti: str = None) -> List[Dict]:
    """추천 시스템 (SQL 최적화: 단일 쿼리로 N+1 문제 해결)"""
    conn = get_connection()
    
    # 1. 기본 쿼리 틀 (참가자 정보 + 방문 통계)
    # LEFT JOIN을 써서 방문 기록이 없는 사람(0회)도 조회되도록 함
    sql = """
        SELECT 
            p.name, p.birth_date, p.gender, p.job, p.mbti, p.phone, p.location, p.signup_route, p.memo,
            COUNT(a.session_id) as visit_count,
            MAX(s.session_date) as last_visit
        FROM participants p
        LEFT JOIN attendance a ON p.name = a.participant_name AND p.birth_date = a.participant_birth
        LEFT JOIN sessions s ON a.session_id = s.session_id
        WHERE p.gender = %s
    """
    params = [gender]

    # 2. 동적 필터 조건 추가 (나이, MBTI)
    if age_min or age_max:
        curr_year = datetime.now().year
        if age_min:
            params.append(curr_year - age_min)
            sql += " AND CAST(SUBSTRING(p.birth_date, 1, 4) AS INTEGER) <= %s"
        if age_max:
            params.append(curr_year - age_max)
            sql += " AND CAST(SUBSTRING(p.birth_date, 1, 4) AS INTEGER) >= %s"
    
    if mbti:
        params.append(f"%{mbti}%")
        sql += " AND p.mbti LIKE %s"

    # 3. [핵심] 제외 로직 (NOT EXISTS 서브쿼리 활용)
    # (1) 현재 세션에 이미 있는 사람 제외
    # (2) 현재 세션 멤버들과 '과거에 만난 적 있는' 사람 제외
    
    # 쿼리에 session_id가 3번 들어갑니다. (현재 멤버 조회용 2번 + 자기 자신 세션 제외용 1번)
    params.extend([session_id, session_id, session_id])
    
    sql += """
        AND NOT EXISTS (
            -- 1. 이미 이번 회차에 등록된 사람 제외
            SELECT 1 FROM attendance curr
            WHERE curr.session_id = %s
            AND curr.participant_name = p.name 
            AND curr.participant_birth = p.birth_date
        )
        AND NOT EXISTS (
            -- 2. 이번 회차 멤버들과 '만난 적 있는' 사람 제외 (겹지인 필터링)
            SELECT 1
            FROM attendance my_history               -- 후보자의 과거 기록
            JOIN attendance met_history              -- 같은 회차였던 사람들 기록
              ON my_history.session_id = met_history.session_id
            JOIN attendance current_session_members  -- 그 사람이 이번 회차 멤버인지 확인
              ON met_history.participant_name = current_session_members.participant_name
              AND met_history.participant_birth = current_session_members.participant_birth
            WHERE my_history.participant_name = p.name
              AND my_history.participant_birth = p.birth_date
              AND current_session_members.session_id = %s  -- 기준: 이번 회차 멤버들
              AND my_history.session_id != %s              -- (혹시 모를 현재 회차 중복 계산 방지)
        )
        GROUP BY p.name, p.birth_date, p.gender, p.job, p.mbti, p.phone, p.location, p.signup_route, p.memo
    """

    # 4. 실행 및 결과 반환
    with get_cursor(conn) as cursor:
        cursor.execute(sql, tuple(params))
        recommendations = [dict(row) for row in cursor.fetchall()]
        
    return recommendations

# ---------------------------------------------------------
# 5. 수정/삭제/엑셀 (Utility)
# ---------------------------------------------------------

def update_participant_memo(name: str, birth_date: str, memo: str):
    """메모 수정"""
    conn = get_connection()
    try:
        with get_cursor(conn) as cursor:
            cursor.execute("UPDATE participants SET memo = %s WHERE name = %s AND birth_date = %s", (memo, name, birth_date))
            conn.commit()
            clear_cache()
    except Exception as e:
        conn.rollback()
        st.error(f"메모 수정 실패: {e}")

def delete_session(session_id: int):
    """회차 삭제 (관련 기록 전체 삭제)"""
    conn = get_connection()
    try:
        with get_cursor(conn) as cursor:
            cursor.execute("SELECT DISTINCT participant_name, participant_birth FROM attendance WHERE session_id = %s", (session_id,))
            participants_in_session = cursor.fetchall()
            
            cursor.execute("DELETE FROM attendance WHERE session_id = %s", (session_id,))
            cursor.execute("DELETE FROM sessions WHERE session_id = %s", (session_id,))
            
            # 고아 참가자 삭제
            for p in participants_in_session:
                cursor.execute("SELECT 1 FROM attendance WHERE participant_name = %s AND participant_birth = %s LIMIT 1", 
                               (p['participant_name'], p['participant_birth']))
                if not cursor.fetchone():
                    cursor.execute("DELETE FROM participants WHERE name = %s AND birth_date = %s", 
                                   (p['participant_name'], p['participant_birth']))
            
            conn.commit()
            clear_cache()
            print(f"✅ {session_id}회차 삭제 완료!")
    except Exception as e:
        conn.rollback()
        raise e

def remove_participant_from_session(session_id: int, participant_name: str, participant_birth: str):
    """특정 회차에서 참가자 제거 + 방문 이력 없으면 DB에서 완전 삭제 (고아 제거)"""
    conn = get_connection()
    try:
        with get_cursor(conn) as cursor:
            # 1. 이번 회차 출석 기록 삭제
            cursor.execute("""
                DELETE FROM attendance 
                WHERE session_id = %s AND participant_name = %s AND participant_birth = %s
            """, (session_id, participant_name, participant_birth))
            
            # 2. [핵심] 남은 방문 이력이 있는지 확인
            cursor.execute("""
                SELECT 1 FROM attendance 
                WHERE participant_name = %s AND participant_birth = %s 
                LIMIT 1
            """, (participant_name, participant_birth))
            
            # 3. 이력이 하나도 없으면 -> 참가자 DB에서도 완전 삭제
            if not cursor.fetchone():
                cursor.execute("""
                    DELETE FROM participants 
                    WHERE name = %s AND birth_date = %s
                """, (participant_name, participant_birth))
                print(f"🧹 {participant_name}님 방문 기록 0회 -> DB에서 자동 삭제됨")

            conn.commit()
            clear_cache()
            print(f"✅ {participant_name} 제거 완료!")
    except Exception as e:
        conn.rollback()
        raise e

def delete_participant(participant_name: str, participant_birth: str):
    """참가자 완전 삭제"""
    conn = get_connection()
    try:
        with get_cursor(conn) as cursor:
            cursor.execute("DELETE FROM attendance WHERE participant_name = %s AND participant_birth = %s", (participant_name, participant_birth))
            cursor.execute("DELETE FROM participants WHERE name = %s AND birth_date = %s", (participant_name, participant_birth))
            conn.commit()
            clear_cache()
            print(f"✅ {participant_name} 삭제 완료!")
    except Exception as e:
        conn.rollback()
        raise e

def import_excel_file(file_path):
    """엑셀 파일 임포트 (최적화)"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    conn = get_connection()
    total = 0
    try:
        with get_cursor(conn) as cursor:
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                s_name_clean = sheet_name.replace("의 사본", "").strip()
                print(f"Processing: {s_name_clean}")
                
                match = re.search(r'(\d{4})(\d{2})(\d{2})', s_name_clean)
                if not match: continue
                s_date = f"{match.group(1)}-{match.group(2)}-{match.group(3)}"
                
                a1 = str(sheet['A1'].value).strip() if sheet['A1'].value else ""
                host = str(sheet['N2'].value).strip() if sheet['N2'].value else "미정"
                
                #s_time = "미정"
                #t_match = re.search(r'(\d{1,2}):(\d{2})\s*(AM|PM)', a1, re.IGNORECASE)
                #if t_match:
                #    h, m, mer = int(t_match.group(1)), int(t_match.group(2)), t_match.group(3).upper()
                #    if mer == 'PM' and h != 12: h += 12
                #    elif mer == 'AM' and h == 12: h = 0
                #    s_time = f"{h:02d}:{m:02d}"

                s_time = "미정"
                # (\d{1,2}) : 시간 (1~2자리)
                # (?::(\d{2}))? : 콜론과 분은 '있을 수도 있고 없을 수도 있음' (?)
                # \s* : 공백 허용
                # (AM|PM) : 오전/오후 필수
                t_match = re.search(r'(\d{1,2})(?::(\d{2}))?\s*(AM|PM)', a1, re.IGNORECASE)
                
                if t_match:
                    h = int(t_match.group(1))
                    # 분(group 2)이 없으면 0분으로 처리
                    m = int(t_match.group(2)) if t_match.group(2) else 0
                    mer = t_match.group(3).upper()
                    
                    if mer == 'PM' and h != 12: h += 12
                    elif mer == 'AM' and h == 12: h = 0
                    
                    s_time = f"{h:02d}:{m:02d}"
                
                theme_match = re.search(r'-\s*(.+)$', a1)
                theme = theme_match.group(1).strip() if theme_match else a1
                
                cursor.execute("INSERT INTO sessions (session_date, session_time, theme, host) VALUES (%s, %s, %s, %s) RETURNING session_id", 
                               (s_date, s_time, theme, host))
                sid = cursor.fetchone()['session_id']
                
                cnt = 0
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    try:
                        vals = [str(c).strip() if c else "" for c in row]
                        if len(vals) < 12: continue
                        gender, nick, name, phone, _, _, loc, birth, job, mbti, intro, route = vals[:12]
                        if not name or not birth or birth == "-": continue
                        
                        g_code = 'M' if gender.upper() in ['M', '남', '남자', '男'] else 'F'
                        b_clean = re.sub(r'\D', '', birth)
                        if len(b_clean) != 4: continue
                        b_date = f"{b_clean}-01-01"
                        p_clean = re.sub(r'\D', '', phone)
                        
                        cursor.execute("""
                            INSERT INTO participants (name, birth_date, gender, nickname, phone, location, job, mbti, intro, signup_route, first_visit_date)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) ON CONFLICT (name, birth_date) DO NOTHING
                        """, (name, b_date, g_code, nick, p_clean, loc, job, mbti, intro, route, s_date))
                        
                        cursor.execute("INSERT INTO attendance (session_id, participant_name, participant_birth) VALUES (%s, %s, %s) ON CONFLICT DO NOTHING", 
                                       (sid, name, b_date))
                        cnt += 1
                    except: continue
                total += cnt
                print(f" -> {s_date}: {cnt}명")
            conn.commit()
    except Exception as e:
        conn.rollback()
        st.error(f"엑셀 임포트 오류: {e}")
    clear_cache()
    print(f"🎉 임포트 완료! 총 {total}명")